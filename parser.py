from __future__ import annotations

import io
import json
import re
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import pandas as pd
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# -----------------------------
# Static rules / normalizations
# -----------------------------

DEFAULT_BRAND_ALIASES: Dict[str, str] = {
    "TOY": "TOYOTA",
    "TOY.": "TOYOTA",
    "TOYO": "TOYOTA",
    "TOYOTA": "TOYOTA",
    "MIT": "MITSUBISHI",
    "MIT.": "MITSUBISHI",
    "MITS": "MITSUBISHI",
    "MITS.": "MITSUBISHI",
    "MITSUBISHI": "MITSUBISHI",
    "MITSUBUSHI": "MITSUBISHI",
    "NIS": "NISSAN",
    "NIS.": "NISSAN",
    "NISSAN": "NISSAN",
    "HON": "HONDA",
    "HON.": "HONDA",
    "HONDA": "HONDA",
    "ISU": "ISUZU",
    "ISU.": "ISUZU",
    "ISUZU": "ISUZU",
    "HYU": "HYUNDAI",
    "HYU.": "HYUNDAI",
    "HYUNDAI": "HYUNDAI",
    "MAZ": "MAZDA",
    "MAZ.": "MAZDA",
    "MAZDA": "MAZDA",
    "SUZ": "SUZUKI",
    "SUZ.": "SUZUKI",
    "SUZUKI": "SUZUKI",
    "KIA": "KIA",
    "FORD": "FORD",
    "CHEV": "CHEVROLET",
    "CHEV.": "CHEVROLET",
    "CHEVROLET": "CHEVROLET",
    "DAIHATSU": "DAIHATSU",
    "MERCEDES": "MERCEDES BENZ",
    "BENZ": "MERCEDES BENZ",
    "MERCEDES BENZ": "MERCEDES BENZ",
    "SUBARU": "SUBARU",
    "PROTON": "PROTON",
    "FOTON": "FOTON",
    "DAEWOO": "DAEWOO",
    "MAHINDRA": "MAHINDRA",
    "ASIATOPIC": "ASIATOPIC",
    "NUVO-PRO": "NUVO-PRO",
}

DEFAULT_ENGINE_CODES = {
    "18R", "2E", "3K", "4K", "12R", "2C", "2L", "3L", "4D31", "4D32", "4DR5", "4HF1",
    "4BC2", "4BC1", "4BB1", "4BA1", "4BD1", "4BE1", "4M40", "4G63", "4D56", "4JA1", "4JB1",
    "4JG2", "4JG2", "4JB1", "1KD", "2KD", "K6A", "C190", "C240", "4BA1", "4BC2", "4HF1",
    "4HF1", "4JG2", "4JA1", "4JB1", "4D56", "4M40", "4G64", "4G15", "4G13", "4G92",
}

POSITION_ALIASES = {
    "axle": {
        "FR": "Front",
        "FRT": "Front",
        "FRT.": "Front",
        "FRONT": "Front",
        "RR": "Rear",
        "REAR": "Rear",
        "FRT.& RR": "Front & Rear",
        "FRT. & RR": "Front & Rear",
        "FRT.&RR": "Front & Rear",
        "FRT & RR": "Front & Rear",
        "FRONT & REAR": "Front & Rear",
    },
    "side": {
        "LH": "Left",
        "LEFT": "Left",
        "RH": "Right",
        "RIGHT": "Right",
    },
    "vertical": {
        "LOW": "Lower",
        "LOWER": "Lower",
        "UP": "Upper",
        "UPPER": "Upper",
    },
    "mount": {
        "INNER": "Inner",
        "OUTER": "Outer",
    },
}

PROTECTED_MODEL_PATTERNS = [
    r"\bBONGO\s+2000\b",
    r"\bE2000\b",
    r"\bKC2700\b",
    r"\bCUBE\s+1\.5\b",
    r"\bMIRAGE\s+G4\b",
    r"\bMU-X\s+13\b",
    r"\bL300\b",
    r"\bBB\s+01['\"]?-11['\"]?\b",
]

QUALIFIER_WORDS = {
    "ORIG", "TYPE", "LEC", "SERIES", "DOUBLE", "TIRE", "GAS", "DSL", "SURPLUS", "D.T.", "D.T", "BB",
    "VAN", "COMMUTER", "FX", "XLE", "GLI", "NHR", "NKR", "NPR", "DSL.", "GAS.", "PV", "MM",
}

ENGINE_IGNORE_TOKENS = {
    "4X2", "4X4", "2WD", "4WD", "12V", "16V", "35MM", "58MM", "66MM", "75MM", "76MM", "30", "30.",
    "58", "66", "70", "76", "85MM", "90MM", "95MM", "100MM", "2000", "2012", "2014", "2016",
}

MAIN_COLUMNS = [
    "Source Line",
    "Supplier Brand",
    "Category",
    "Code",
    "Axle",
    "Side",
    "Vertical",
    "Mount",
    "Brand",
    "Model",
    "Engine",
    "Year",
    "Original Price (PHP)",
    "Your Price (PHP)",
    "Use Price (PHP)",
    "Page",
    "Confidence",
    "Review Status",
    "Pattern Notes",
]

RAW_COLUMNS = [
    "Page",
    "Supplier Brand",
    "Category",
    "Code",
    "Axle",
    "Description",
    "Original Price (PHP)",
    "Extraction Notes",
]

CONFLICT_COLUMNS = [
    "Row Key",
    "Page",
    "Code",
    "Field",
    "Severity",
    "Issue",
    "Chosen Value",
    "Alternatives",
]

EVIDENCE_COLUMNS = [
    "Row Key",
    "Page",
    "Code",
    "Field",
    "Value",
    "Confidence",
    "Source",
    "Reason",
]

NOISE_TERMS = [
    "TIHLUCK",
    "PASAY",
    "EMAIL",
    "UPDATE SEPT",
    "OF 18",
    "(02)",
    "GUARANTEED",
    "TRADING",
    "CORPORATION",
    "GMAIL.COM",
]

CODE_REGEX = re.compile(r"^(?P<code>(?:VAX|VA|VKX|VK)-[A-Z0-9*]+)\b", re.IGNORECASE)
PRICE_TOKEN_REGEX = re.compile(r"^\d{1,3}(?:,\d{3})*(?:\.\d{2})$")
LINE_PRICE_REGEX = re.compile(r"(?P<price>\d{1,3}(?:,\d{3})*(?:\.\d{2}))\s*$")
SPACED_PRICE_REGEX = re.compile(r"(?P<price>(?:\d\s*){1,4}(?:,\s*(?:\d\s*){3})?\.\s*\d\s*\d)\s*$")
GENERIC_ENGINE_REGEX = re.compile(r"^(?:[1-9][A-Z]{1,3}\d{1,3}[A-Z]?|\d{1,2}[A-Z]{1,3}|C-?\d{3})$", re.IGNORECASE)
YEAR_RANGE_4D = re.compile(r"(?P<all>(?P<s>19\d{2}|20\d{2})\s*(?:-|TO)\s*(?P<e>19\d{2}|20\d{2}|UP))", re.IGNORECASE)
YEAR_RANGE_2D = re.compile(r"(?P<all>[\"']?(?P<s>\d{2})[\"']?\s*-\s*[\"']?(?P<e>\d{2}|UP)[\"']?)", re.IGNORECASE)
YEAR_SINGLE_4D = re.compile(r"(?P<all>19\d{2}|20\d{2})", re.IGNORECASE)
YEAR_SINGLE_QUOTED = re.compile(r"(?P<all>[\"']?(?P<y>\d{2})[\"'])", re.IGNORECASE)
YEAR_SINGLE_TRAILING = re.compile(r"(?P<all>(?<![A-Z0-9])(?P<y>\d{2})(?![A-Z0-9]))$", re.IGNORECASE)


@dataclass
class CatalogRow:
    page: int
    supplier_brand: str
    category: str
    code: str
    axle: str
    description: str
    price: Optional[float]
    source_line: str
    extraction_notes: str = ""


@dataclass
class ParsedCatalog:
    raw_rows: List[CatalogRow]
    application_rows: List[Dict[str, object]]
    review_rows: List[Dict[str, object]]
    warnings: List[str]
    conflict_rows: List[Dict[str, object]] = field(default_factory=list)
    evidence_rows: List[Dict[str, object]] = field(default_factory=list)
    metrics: Dict[str, int] = field(default_factory=dict)


# -----------------------------
# Context / dictionaries
# -----------------------------


def _merge_aliases(custom_aliases: Optional[Dict[str, str]] = None) -> Dict[str, str]:
    aliases = {k.upper().strip(): v.upper().strip() for k, v in DEFAULT_BRAND_ALIASES.items()}
    for key, value in (custom_aliases or {}).items():
        if str(key).strip() and str(value).strip():
            aliases[str(key).upper().strip()] = str(value).upper().strip()
    return aliases


def _brand_regex(aliases: Dict[str, str]) -> re.Pattern[str]:
    keys = sorted(aliases.keys(), key=len, reverse=True)
    return re.compile(r"^(?P<brand>" + "|".join(re.escape(k) for k in keys) + r")\b[\s.-]*(?P<rest>.*)$", re.IGNORECASE)


def _protected_patterns(extra_phrases: Optional[List[str]] = None) -> List[re.Pattern[str]]:
    patterns = [re.compile(p, re.IGNORECASE) for p in PROTECTED_MODEL_PATTERNS]
    for phrase in extra_phrases or []:
        phrase = str(phrase).strip()
        if phrase:
            patterns.append(re.compile(re.escape(phrase), re.IGNORECASE))
    return patterns




def _starts_with_brand(text: str, aliases: Optional[Dict[str, str]] = None) -> bool:
    alias_map = aliases or _merge_aliases()
    pattern = _brand_regex(alias_map)
    return bool(pattern.match(_clean_text(text)))


# -----------------------------
# Text helpers
# -----------------------------


def _clean_text(text: str) -> str:
    value = str(text or "")
    value = (
        value.replace("–", "-")
        .replace("—", "-")
        .replace("’", "'")
        .replace("“", '"')
        .replace("”", '"')
        .replace("½", "1/2")
    )
    value = re.sub(r"(?<=\d)\s*,\s*(?=\d{3}\b)", ",", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def _cleanup_model(text: str) -> str:
    value = _clean_text(text)
    value = value.replace(".ORIG", " ORIG")
    value = value.replace("HI ACE", "HIACE")
    value = value.replace("STA FE", "SANTA FE")
    value = value.replace("I1O", "I10")
    value = re.sub(r"\b\d{1,3}(?:,\d{3})*(?:\.\d{2})\b$", "", value).strip()
    value = value.replace("  ", " ")
    value = value.strip(" ,-/")
    return re.sub(r"\s+", " ", value)


def _model_key(model: str) -> str:
    return re.sub(r"[^A-Z0-9]+", " ", _cleanup_model(model).upper()).strip()


def _safe_float(value: str | None) -> Optional[float]:
    if value is None:
        return None
    text = _clean_text(value).replace(",", "").replace(" ", "")
    try:
        return float(text)
    except Exception:
        return None


def _extract_price_tail(text: str) -> Tuple[Optional[float], str, str]:
    raw = _clean_text(text)
    m = LINE_PRICE_REGEX.search(raw)
    if m:
        price_text = m.group("price")
        return _safe_float(price_text), _clean_text(raw[: m.start()]), price_text
    return None, raw, ""


def _expand_short_year(two_digit: str) -> int:
    year = int(two_digit)
    return 2000 + year if year <= 35 else 1900 + year


def _normalize_year(value: str) -> str:
    text = _clean_text(value).replace('\"', "'").upper().replace(" TO ", "-")
    text = text.replace(" - UP", "-UP").replace("- UP", "-UP")
    text = text.replace("''", "'")

    m = re.fullmatch(r"[']?([0-9]{2})[']?\s*-\s*[']?([0-9]{2})[']?", text)
    if m:
        return f"{_expand_short_year(m.group(1))}-{_expand_short_year(m.group(2))}"

    m = re.fullmatch(r"[']?([0-9]{2})[']?\s*-\s*UP", text)
    if m:
        return f"{_expand_short_year(m.group(1))}-Up"

    m = re.fullmatch(r"([0-9]{4})\s*-\s*([0-9]{4})", text)
    if m:
        return f"{m.group(1)}-{m.group(2)}"

    m = re.fullmatch(r"([0-9]{4})\s*-\s*UP", text)
    if m:
        return f"{m.group(1)}-Up"

    m = re.fullmatch(r"[']?([0-9]{2})[']?", text)
    if m:
        return str(_expand_short_year(m.group(1)))

    m = re.fullmatch(r"([0-9]{4})", text)
    if m:
        return m.group(1)

    return value


def _extract_year(text: str, protected_patterns: List[re.Pattern[str]]) -> Tuple[str, str, str, float]:
    raw = _clean_text(text)
    protected_spans: List[Tuple[int, int]] = []
    for pattern in protected_patterns:
        for match in pattern.finditer(raw):
            protected_spans.append(match.span())

    candidates: List[Tuple[int, int, str, str, float]] = []
    patterns = [
        (YEAR_RANGE_4D, 0.96, "four-digit year range"),
        (YEAR_RANGE_2D, 0.93, "short year range"),
        (YEAR_SINGLE_4D, 0.86, "single four-digit year"),
        (YEAR_SINGLE_QUOTED, 0.91, "quoted short year"),
        (YEAR_SINGLE_TRAILING, 0.78, "trailing short year"),
    ]
    for pattern, score, label in patterns:
        for match in pattern.finditer(raw):
            span = match.span("all") if "all" in match.groupdict() else match.span()
            if any(span[0] >= s0 and span[1] <= s1 for s0, s1 in protected_spans):
                continue
            matched_text = match.group("all") if "all" in match.groupdict() else match.group(0)
            candidates.append((span[0], span[1], matched_text, label, score))

    if not candidates:
        return "", raw, "", 0.0

    # Prefer later matches for year-at-end patterns, then longer spans.
    candidates.sort(key=lambda item: (item[1] - item[0], item[0]))
    start, end, matched_text, label, score = candidates[-1]
    year = _normalize_year(matched_text)
    remaining = _cleanup_model((raw[:start] + " " + raw[end:]).strip())
    return year, remaining, label, score


def _detect_engine_codes(text: str) -> List[str]:
    raw = _clean_text(text).upper().replace("C-240", "C240").replace("C-190", "C190")
    found: List[str] = []
    tokens = re.findall(r"[A-Z0-9.-]+", raw)
    for token in tokens:
        token = token.strip(".").replace("C-240", "C240").replace("C-190", "C190")
        if not token or token in ENGINE_IGNORE_TOKENS:
            continue
        if token in DEFAULT_ENGINE_CODES:
            if token not in found:
                found.append(token)
            continue
        if GENERIC_ENGINE_REGEX.fullmatch(token):
            if "X" in token or token.endswith("MM") or token.endswith("V") or token in ENGINE_IGNORE_TOKENS:
                continue
            if token not in found:
                found.append(token)
    return found


def _normalize_axis_value(value: str, axis: str) -> str:
    text = _clean_text(value).upper()
    mapping = POSITION_ALIASES[axis]
    return mapping.get(text, value)


# -----------------------------
# Layout-aware PDF reading
# -----------------------------


def _parse_line_text(text: str) -> Optional[Dict[str, object]]:
    raw = _clean_text(text)
    if not raw:
        return None
    m_code = CODE_REGEX.match(raw)
    if not m_code:
        return None
    code = m_code.group("code").upper()
    remainder = _clean_text(raw[m_code.end():])

    price, remainder, price_text = _extract_price_tail(remainder)

    axle = ""
    axle_patterns = [
        (r"^(FRT\.?\s*&\s*RR|FRT\.?&\s*RR|FRONT\s*&\s*REAR)\b", "Front & Rear"),
        (r"^(REAR)\b", "Rear"),
        (r"^(FRONT|FRT\.?)\b", "Front"),
        (r"^(RR)\b", "Rear"),
        (r"^(FR)\b", "Front"),
    ]
    for pattern, normalized in axle_patterns:
        m_axle = re.match(pattern, remainder, re.IGNORECASE)
        if m_axle:
            axle = normalized
            remainder = _clean_text(remainder[m_axle.end():])
            break

    return {
        "code": code,
        "axle": axle,
        "description": remainder,
        "price": price,
        "extraction_notes": "layout-aware row detection; text-first line parse",
    }


def _group_page_lines(page) -> List[Dict[str, object]]:
    words = page.extract_words(use_text_flow=False, keep_blank_chars=False, x_tolerance=1.2, y_tolerance=2.0)
    grouped: List[Dict[str, object]] = []
    for word in sorted(words, key=lambda x: (round(x["top"], 1), x["x0"])):
        placed = False
        for line in grouped:
            if abs(float(line["top"]) - float(word["top"])) <= 2.3:
                line["words"].append(word)
                placed = True
                break
        if not placed:
            grouped.append({"top": float(word["top"]), "words": [word]})

    output: List[Dict[str, object]] = []
    for line in sorted(grouped, key=lambda x: x["top"]):
        row_words = sorted(line["words"], key=lambda x: x["x0"])
        output.append(
            {
                "top": line["top"],
                "min_x": min(word["x0"] for word in row_words),
                "max_x": max(word["x1"] for word in row_words),
                "words": row_words,
                "text": _clean_text(" ".join(word["text"] for word in row_words)),
            }
        )
    return output


def _extract_row_from_line(line: Dict[str, object], page_width: float) -> Optional[Dict[str, object]]:
    words = line["words"]
    parsed = _parse_line_text(str(line.get("text", "")))
    if parsed:
        right_side_tokens = [_clean_text(w["text"]) for w in words if w["x0"] > page_width * 0.82]
        coord_price_text = "".join(right_side_tokens)
        coord_price = _safe_float(coord_price_text) if coord_price_text and PRICE_TOKEN_REGEX.match(coord_price_text) else None
        if parsed["price"] is None and coord_price is not None:
            parsed["price"] = coord_price
            parsed["description"] = _clean_text(" ".join(w["text"] for w in words if page_width * 0.20 <= w["x0"] < page_width * 0.82))
            parsed["extraction_notes"] = str(parsed["extraction_notes"]) + "; price recovered from right-edge coordinates"
        elif coord_price is not None and parsed["price"] is not None and parsed["price"] > coord_price * 2:
            parsed["price"] = coord_price
            parsed["description"] = _clean_text(" ".join(w["text"] for w in words if page_width * 0.20 <= w["x0"] < page_width * 0.82))
            parsed["extraction_notes"] = str(parsed["extraction_notes"]) + "; suspicious text price replaced by right-edge coordinates"
        if parsed["price"] is None:
            parsed["extraction_notes"] = str(parsed["extraction_notes"]) + "; price missing on detected row"
        return parsed

    code = ""
    right_side_tokens: List[str] = []
    for word in words:
        token = _clean_text(word["text"])
        if not code and CODE_REGEX.match(token) and word["x0"] < page_width * 0.20:
            code = CODE_REGEX.match(token)["code"].upper()
        if word["x0"] > page_width * 0.82:
            right_side_tokens.append(token)

    if not code:
        return None

    price_text = "".join(right_side_tokens)
    price = _safe_float(price_text) if price_text else None

    axle_words = [w for w in words if page_width * 0.13 <= w["x0"] < page_width * 0.26]
    desc_words = [w for w in words if page_width * 0.20 <= w["x0"] < page_width * 0.82]

    extraction_notes = ["layout-aware row detection; coordinate fallback"]
    if price_text and price is None:
        extraction_notes.append(f"price token normalized from '{price_text}'")
    if price is None:
        extraction_notes.append("price missing on detected row")

    description = _clean_text(" ".join(w["text"] for w in desc_words))
    if price_text and description.endswith(price_text):
        description = _clean_text(description[: -len(price_text)])

    return {
        "code": code,
        "axle": _clean_text(" ".join(w["text"] for w in axle_words)),
        "description": description,
        "price": price,
        "extraction_notes": "; ".join(extraction_notes),
    }


def extract_catalog_rows(pdf_path: str | Path) -> Tuple[List[CatalogRow], List[str]]:
    rows: List[CatalogRow] = []
    warnings: List[str] = []
    pdf_name = Path(pdf_path).stem.upper()
    supplier_brand = "NUVO-PRO" if "NUVO" in pdf_name else ""
    category = ""
    alias_map = _merge_aliases()

    with pdfplumber.open(str(pdf_path)) as pdf:
        for page_number, page in enumerate(pdf.pages, start=1):
            current: Optional[Dict[str, object]] = None
            pending_prefix_lines: List[str] = []
            page_width = float(page.width)
            desc_min_x = page_width * 0.20
            desc_max_x = page_width * 0.87

            def _finalize_current() -> None:
                nonlocal current
                if current:
                    rows.append(
                        CatalogRow(
                            page=page_number,
                            supplier_brand=supplier_brand or "",
                            category=category or "",
                            code=str(current["code"]),
                            axle=str(current["axle"]),
                            description=str(current["description"]),
                            price=current["price"],
                            source_line=_clean_text(f"{current['code']} {current['axle']} {current['description']}"),
                            extraction_notes=str(current["extraction_notes"]),
                        )
                    )
                    current = None

            for line in _group_page_lines(page):
                line_text = line["text"]
                upper = line_text.upper()
                if not line_text:
                    continue

                if "NUVO-PRO" in upper and not supplier_brand:
                    supplier_brand = "NUVO-PRO"
                if "BRAKE PAD" in upper:
                    category = "Brake Pad"
                    current = None
                    pending_prefix_lines = []
                    continue
                if "BRAKE SHOE" in upper:
                    category = "Brake Shoe"
                    current = None
                    pending_prefix_lines = []
                    continue
                if re.fullmatch(r"\d+\s+OF\s+\d+", upper):
                    continue
                if any(noise in upper for noise in NOISE_TERMS):
                    continue

                extracted = _extract_row_from_line(line, page_width)
                if extracted:
                    if pending_prefix_lines:
                        pending_text = _clean_text(" ".join(pending_prefix_lines))
                        if pending_text:
                            if current and _starts_with_brand(str(extracted.get("description", "")), alias_map):
                                current["description"] = _clean_text(f"{current['description']} {pending_text}")
                                current["extraction_notes"] = f"{current['extraction_notes']}; wrapped description continuation joined"
                            else:
                                extracted["description"] = _clean_text(f"{pending_text} {extracted['description']}")
                                extracted["extraction_notes"] = str(extracted["extraction_notes"]) + "; wrapped description prefix joined"
                        pending_prefix_lines = []

                    _finalize_current()
                    current = extracted
                    continue

                in_description_band = desc_min_x <= line["min_x"] < desc_max_x
                if current and in_description_band:
                    if current.get("price") is None:
                        current["description"] = _clean_text(f"{current['description']} {line_text}")
                        current["extraction_notes"] = f"{current['extraction_notes']}; continuation joined"
                    else:
                        pending_prefix_lines.append(line_text)
                    continue

                if (not current) and in_description_band:
                    pending_prefix_lines.append(line_text)

            if pending_prefix_lines and current:
                pending_text = _clean_text(" ".join(pending_prefix_lines))
                if pending_text:
                    current["description"] = _clean_text(f"{current['description']} {pending_text}")
                    current["extraction_notes"] = f"{current['extraction_notes']}; trailing wrapped continuation joined"
            _finalize_current()

    if not rows:
        warnings.append("No catalog rows were extracted. This PDF may be image-only or use a very different layout.")
    if any(row.price is None for row in rows):
        warnings.append("Some rows were extracted without a price. Review Catalog Rows for missing values.")
    return rows, warnings


# -----------------------------
# Segment parsing / normalization
# -----------------------------


def _normalize_axle(value: str) -> str:
    text = _clean_text(value).upper()
    if text in POSITION_ALIASES["axle"]:
        return POSITION_ALIASES["axle"][text]
    if text in {"FRONT", "REAR"}:
        return text.title()
    if text in {"FR", "FRT", "FRT."}:
        return "Front"
    if text == "RR":
        return "Rear"
    return _clean_text(value).title()


def _tokenize_slash_segments(text: str) -> List[str]:
    raw = _clean_text(text)
    raw = re.sub(r"(?<=[A-Z0-9'\"])/(?=[A-Z0-9])", " / ", raw)
    parts = [part.strip(" ,-") for part in re.split(r"\s*/\s*", raw) if part.strip(" ,-")]
    return parts


def _looks_like_engine_only_fragment(fragment: str) -> bool:
    raw = _clean_text(fragment).upper()
    if not raw:
        return False
    tokens = [re.sub(r"[^A-Z0-9.-]", "", t) for t in raw.replace("(", " ").replace(")", " ").split() if t]
    engine_like = 0
    for token in tokens:
        token = token.replace("C-240", "C240")
        if token in DEFAULT_ENGINE_CODES or GENERIC_ENGINE_REGEX.fullmatch(token):
            engine_like += 1
        elif token in QUALIFIER_WORDS or re.fullmatch(r"\d+(?:MM|\"|IN|/\d+)?", token) or token in {"&", "X", "2WD", "4WD", "4X2", "4X4"}:
            continue
        else:
            return False
    return engine_like > 0


def _looks_like_attach_previous(fragment: str, aliases: Dict[str, str]) -> bool:
    raw = _clean_text(fragment).upper()
    if not raw:
        return False
    if _looks_like_engine_only_fragment(raw):
        return True
    first = re.sub(r"[^A-Z0-9.-]", "", raw.split()[0])
    if first in QUALIFIER_WORDS:
        return True
    if re.fullmatch(r"[\"']?\d{2}(?:\s*-\s*[\"']?\d{2}|[\"'])?", raw):
        return True
    if first in {"12V", "16V", "4X2", "4X4", "2WD", "4WD"}:
        return True
    if first in aliases:
        return False
    return False


def _split_segments(description: str, aliases: Dict[str, str]) -> Tuple[List[str], List[str]]:
    raw_parts = _tokenize_slash_segments(description)
    if not raw_parts:
        return [], []

    segment_records: List[Dict[str, object]] = []
    split_notes: List[str] = []
    for fragment in raw_parts:
        if not segment_records:
            segment_records.append({"base": fragment, "suffixes": []})
            continue

        if _looks_like_engine_only_fragment(fragment):
            if len(segment_records) > 1:
                for rec in segment_records:
                    rec["suffixes"].append(fragment)
                split_notes.append(f"shared fragment applied to all segments: {fragment}")
            else:
                segment_records[-1]["suffixes"].append(fragment)
                split_notes.append(f"engine fragment attached to previous segment: {fragment}")
            continue

        if _looks_like_attach_previous(fragment, aliases):
            segment_records[-1]["suffixes"].append(fragment)
            split_notes.append(f"qualifier attached to previous segment: {fragment}")
            continue

        segment_records.append({"base": fragment, "suffixes": []})

    segments: List[str] = []
    for rec in segment_records:
        text = _clean_text(str(rec["base"]))
        if rec["suffixes"]:
            text = _clean_text(text + " " + " ".join(str(x) for x in rec["suffixes"]))
        segments.append(text)
    return segments, split_notes


def _extract_position_tokens(text: str) -> Tuple[Dict[str, str], str, List[str], Dict[str, float]]:
    raw = _clean_text(text)
    evidence: List[str] = []
    field_scores: Dict[str, float] = {}
    result = {"side": "", "vertical": "", "mount": "", "axle": ""}

    tokens = raw.split()
    kept: List[str] = []
    for token in tokens:
        normalized = re.sub(r"[^A-Z&.-]", "", token.upper())
        matched = False
        for axis in ["axle", "side", "vertical", "mount"]:
            mapping = POSITION_ALIASES[axis]
            if normalized in mapping:
                result[axis] = mapping[normalized]
                evidence.append(f"{axis} token {normalized}")
                field_scores[axis] = 0.92 if axis != "axle" else 0.96
                matched = True
                break
        if not matched:
            kept.append(token)
    return result, _cleanup_model(" ".join(kept)), evidence, field_scores


def _remove_engine_tokens(text: str, tokens: List[str]) -> str:
    raw = _clean_text(text)
    for token in tokens:
        pattern = re.compile(rf"\b{re.escape(token)}\b", re.IGNORECASE)
        raw = pattern.sub(" ", raw)
    return _cleanup_model(raw)


def _build_model_brand_map(rows: Iterable[Dict[str, object]]) -> Dict[str, str]:
    counter: Counter[Tuple[str, str]] = Counter()
    for row in rows:
        brand = _clean_text(row.get("Brand", "")).upper()
        model = _clean_text(row.get("Model", "")).upper()
        if not brand or not model:
            continue
        first_word = _model_key(model).split(" ")[0] if _model_key(model) else ""
        if first_word:
            counter[(first_word, brand)] += 1

    grouped: Dict[str, List[Tuple[str, int]]] = defaultdict(list)
    for (first_word, brand), count in counter.items():
        grouped[first_word].append((brand, count))

    model_brand: Dict[str, str] = {}
    for first_word, values in grouped.items():
        values.sort(key=lambda x: (-x[1], x[0]))
        model_brand[first_word] = values[0][0]
    return model_brand


def _validate_field(field: str, value: str) -> Tuple[bool, str]:
    text = _clean_text(value)
    if not text:
        return True, ""
    if field == "Year":
        ok = bool(
            re.fullmatch(r"\d{4}", text)
            or re.fullmatch(r"\d{4}-\d{4}", text)
            or re.fullmatch(r"\d{4}-Up", text)
        )
        return ok, "year format" if ok else "invalid year format"
    if field == "Axle":
        return text in {"Front", "Rear", "Front & Rear", ""}, "axle domain"
    if field == "Side":
        return text in {"Left", "Right", ""}, "side domain"
    if field == "Vertical":
        return text in {"Upper", "Lower", ""}, "vertical domain"
    if field == "Mount":
        return text in {"Inner", "Outer", ""}, "mount domain"
    if field == "Engine":
        tokens = [t.strip().upper() for t in text.split(",") if t.strip()]
        ok = all(
            tok in DEFAULT_ENGINE_CODES
            or (GENERIC_ENGINE_REGEX.fullmatch(tok) and tok not in ENGINE_IGNORE_TOKENS and "X" not in tok and not tok.endswith("MM") and not tok.endswith("V"))
            for tok in tokens
        )
        return ok, "engine pattern" if ok else "invalid engine pattern"
    return True, ""


def _parse_segment(
    segment: str,
    carry_brand: str,
    aliases: Dict[str, str],
    protected_patterns: List[re.Pattern[str]],
    row_axle: str,
) -> Tuple[Dict[str, object], List[Dict[str, object]], List[Dict[str, object]]]:
    brand_regex = _brand_regex(aliases)
    raw = _cleanup_model(segment)
    evidence_rows: List[Dict[str, object]] = []
    conflict_rows: List[Dict[str, object]] = []

    field_values = {
        "Brand": "",
        "Model": "",
        "Engine": "",
        "Year": "",
        "Axle": _normalize_axle(row_axle),
        "Side": "",
        "Vertical": "",
        "Mount": "",
    }
    field_conf = {
        "Brand": 0.0,
        "Model": 0.0,
        "Engine": 0.0,
        "Year": 0.0,
        "Axle": 0.96 if row_axle else 0.0,
        "Side": 0.0,
        "Vertical": 0.0,
        "Mount": 0.0,
    }
    notes: List[str] = []

    # Explicit or carried brand.
    m_brand = brand_regex.match(raw)
    if m_brand:
        explicit_brand = aliases[m_brand.group("brand").upper()]
        field_values["Brand"] = explicit_brand
        field_conf["Brand"] = 0.98
        raw = _cleanup_model(m_brand.group("rest"))
        notes.append("explicit brand")
    elif carry_brand:
        field_values["Brand"] = carry_brand
        field_conf["Brand"] = 0.84
        notes.append("brand carried forward")

    # Position tokens from description body.
    positions, raw, pos_notes, pos_scores = _extract_position_tokens(raw)
    for key, value in positions.items():
        mapped_field = key.capitalize() if key != "axle" else "Axle"
        if value:
            field_values[mapped_field] = value
            field_conf[mapped_field] = max(field_conf.get(mapped_field, 0.0), pos_scores.get(key, 0.0))
    notes.extend(pos_notes)

    # Year parsing with protected phrases.
    year, raw_after_year, year_reason, year_score = _extract_year(raw, protected_patterns)
    if year:
        field_values["Year"] = year
        field_conf["Year"] = year_score
        raw = raw_after_year
        notes.append(year_reason)

    # Engine codes.
    engines = _detect_engine_codes(raw)
    if engines:
        field_values["Engine"] = ", ".join(engines)
        field_conf["Engine"] = 0.91
        raw = _remove_engine_tokens(raw, engines)
        notes.append("engine parsed")

    # Final model cleanup.
    field_values["Model"] = _cleanup_model(raw)
    if field_values["Model"]:
        field_conf["Model"] = 0.93
    else:
        conflict_rows.append(
            {
                "Field": "Model",
                "Severity": "High",
                "Issue": "Model became blank after parsing.",
                "Chosen Value": "",
                "Alternatives": segment,
            }
        )

    # Field-level validation.
    for field_name in ["Year", "Axle", "Side", "Vertical", "Mount", "Engine"]:
        ok, reason = _validate_field(field_name, str(field_values[field_name]))
        if not ok:
            conflict_rows.append(
                {
                    "Field": field_name,
                    "Severity": "Medium",
                    "Issue": reason,
                    "Chosen Value": str(field_values[field_name]),
                    "Alternatives": segment,
                }
            )
            field_conf[field_name] = min(field_conf[field_name], 0.35)

    for field_name, value in field_values.items():
        if _clean_text(str(value)):
            evidence_rows.append(
                {
                    "Field": field_name,
                    "Value": value,
                    "Confidence": round(float(field_conf[field_name]), 2),
                    "Source": "rules",
                    "Reason": ", ".join(notes) if notes else "parsed from segment",
                }
            )

    weighted_fields = {
        "Brand": 1.6,
        "Model": 2.0,
        "Year": 1.0,
        "Engine": 0.8,
        "Axle": 0.9,
        "Side": 0.4,
        "Vertical": 0.4,
        "Mount": 0.4,
    }
    numerator = 0.0
    denominator = 0.0
    for field_name, weight in weighted_fields.items():
        if field_name in {"Brand", "Model"} or _clean_text(str(field_values[field_name])):
            numerator += weight * float(field_conf[field_name])
            denominator += weight
    overall = round((numerator / denominator) if denominator else 0.0, 2)

    if not field_values["Brand"]:
        overall = min(overall, 0.68)
    if not field_values["Model"]:
        overall = min(overall, 0.50)
    if any(item["Severity"] == "High" for item in conflict_rows):
        overall = min(overall, 0.62)

    parsed = {
        "Brand": field_values["Brand"],
        "Model": field_values["Model"],
        "Engine": field_values["Engine"],
        "Year": field_values["Year"],
        "Axle": field_values["Axle"],
        "Side": field_values["Side"],
        "Vertical": field_values["Vertical"],
        "Mount": field_values["Mount"],
        "Confidence": overall,
        "Pattern Notes": ", ".join(dict.fromkeys(notes)),
        "_Field Confidence": field_conf,
    }
    return parsed, evidence_rows, conflict_rows


def expand_catalog_rows(
    raw_rows: List[CatalogRow],
    custom_aliases: Optional[Dict[str, str]] = None,
    protected_phrases: Optional[List[str]] = None,
) -> Tuple[List[Dict[str, object]], List[Dict[str, object]], List[Dict[str, object]], List[Dict[str, object]]]:
    aliases = _merge_aliases(custom_aliases)
    protected_patterns = _protected_patterns(protected_phrases)

    expanded: List[Dict[str, object]] = []
    evidence_rows: List[Dict[str, object]] = []
    conflict_rows: List[Dict[str, object]] = []

    for row in raw_rows:
        segments, split_notes = _split_segments(row.description, aliases)
        carry_brand = ""
        for segment_order, segment in enumerate(segments, start=1):
            parsed, seg_evidence, seg_conflicts = _parse_segment(
                segment=segment,
                carry_brand=carry_brand,
                aliases=aliases,
                protected_patterns=protected_patterns,
                row_axle=row.axle,
            )
            if parsed["Brand"]:
                carry_brand = str(parsed["Brand"])

            row_key = f"P{row.page}|{row.code}|{segment_order}"
            app_row = {
                "_Row Key": row_key,
                "Source Line": row.source_line,
                "Supplier Brand": row.supplier_brand,
                "Category": row.category,
                "Code": row.code,
                "Axle": parsed["Axle"],
                "Side": parsed["Side"],
                "Vertical": parsed["Vertical"],
                "Mount": parsed["Mount"],
                "Brand": parsed["Brand"],
                "Model": parsed["Model"],
                "Engine": parsed["Engine"],
                "Year": parsed["Year"],
                "Original Price (PHP)": row.price,
                "Your Price (PHP)": None,
                "Use Price (PHP)": None,
                "Page": row.page,
                "Confidence": parsed["Confidence"],
                "Review Status": "",
                "Pattern Notes": _clean_text(
                    "; ".join([x for x in [parsed["Pattern Notes"], "; ".join(split_notes)] if x])
                ),
                "_Field Confidence": parsed["_Field Confidence"],
                "_Segment": segment,
                "_Segment Order": segment_order,
            }
            expanded.append(app_row)

            for ev in seg_evidence:
                evidence_rows.append(
                    {
                        "Row Key": row_key,
                        "Page": row.page,
                        "Code": row.code,
                        **ev,
                    }
                )
            for con in seg_conflicts:
                conflict_rows.append(
                    {
                        "Row Key": row_key,
                        "Page": row.page,
                        "Code": row.code,
                        **con,
                    }
                )

    # Second pass: infer missing brands from model map.
    model_brand_map = _build_model_brand_map(expanded)
    for row in expanded:
        if not _clean_text(str(row["Brand"])) and _clean_text(str(row["Model"])):
            first_word = _model_key(str(row["Model"])).split(" ")[0]
            inferred_brand = model_brand_map.get(first_word, "")
            if inferred_brand:
                row["Brand"] = inferred_brand
                row["Confidence"] = round(min(float(row["Confidence"]) + 0.12, 0.92), 2)
                row["Pattern Notes"] = _clean_text(f"{row['Pattern Notes']}; brand inferred from model memory ({inferred_brand})")
                evidence_rows.append(
                    {
                        "Row Key": row["_Row Key"],
                        "Page": row["Page"],
                        "Code": row["Code"],
                        "Field": "Brand",
                        "Value": inferred_brand,
                        "Confidence": 0.72,
                        "Source": "supplier memory",
                        "Reason": f"model first word matched prior rows ({first_word})",
                    }
                )
                conflict_rows.append(
                    {
                        "Row Key": row["_Row Key"],
                        "Page": row["Page"],
                        "Code": row["Code"],
                        "Field": "Brand",
                        "Severity": "Low",
                        "Issue": "Brand was not explicit and was inferred from model memory.",
                        "Chosen Value": inferred_brand,
                        "Alternatives": row["_Segment"],
                    }
                )

    review_rows: List[Dict[str, object]] = []
    for row in expanded:
        status = "Auto-Accepted"
        if not _clean_text(str(row["Brand"])) or not _clean_text(str(row["Model"])):
            status = "Needs Review"
            row["Confidence"] = min(float(row["Confidence"]), 0.60)
        elif float(row["Confidence"]) >= 0.92 and "brand inferred" not in str(row["Pattern Notes"]).lower():
            status = "Auto-Accepted"
        elif float(row["Confidence"]) >= 0.78:
            status = "Review Suggested"
        else:
            status = "Needs Review"
        if any(c["Row Key"] == row["_Row Key"] and c["Severity"] == "High" for c in conflict_rows):
            status = "Needs Review"
            row["Confidence"] = min(float(row["Confidence"]), 0.62)

        row["Review Status"] = status
        row["Confidence"] = round(float(row["Confidence"]), 2)
        if status != "Auto-Accepted":
            review_rows.append({column: row.get(column, "") for column in MAIN_COLUMNS})

    expanded.sort(key=lambda x: (int(x["Page"]), str(x["Code"]), int(x["_Segment Order"])))
    return expanded, review_rows, conflict_rows, evidence_rows


def parse_catalog_pdf(
    pdf_path: str | Path,
    custom_aliases: Optional[Dict[str, str]] = None,
    protected_phrases: Optional[List[str]] = None,
) -> ParsedCatalog:
    raw_rows, warnings = extract_catalog_rows(pdf_path)
    application_rows, review_rows, conflict_rows, evidence_rows = expand_catalog_rows(
        raw_rows=raw_rows,
        custom_aliases=custom_aliases,
        protected_phrases=protected_phrases,
    )
    metrics = {
        "catalog_rows": len(raw_rows),
        "application_rows": len(application_rows),
        "auto_accepted": sum(1 for row in application_rows if row["Review Status"] == "Auto-Accepted"),
        "review_suggested": sum(1 for row in application_rows if row["Review Status"] == "Review Suggested"),
        "needs_review": sum(1 for row in application_rows if row["Review Status"] == "Needs Review"),
        "conflicts": len(conflict_rows),
        "evidence_rows": len(evidence_rows),
    }
    return ParsedCatalog(
        raw_rows=raw_rows,
        application_rows=application_rows,
        review_rows=review_rows,
        warnings=warnings,
        conflict_rows=conflict_rows,
        evidence_rows=evidence_rows,
        metrics=metrics,
    )


# -----------------------------
# DataFrame helpers
# -----------------------------


def applications_dataframe(rows: List[Dict[str, object]]) -> pd.DataFrame:
    return pd.DataFrame([{column: row.get(column, "") for column in MAIN_COLUMNS} for row in rows], columns=MAIN_COLUMNS)


def raw_rows_dataframe(rows: List[CatalogRow]) -> pd.DataFrame:
    data = []
    for row in rows:
        data.append(
            {
                "Page": row.page,
                "Supplier Brand": row.supplier_brand,
                "Category": row.category,
                "Code": row.code,
                "Axle": _normalize_axle(row.axle),
                "Description": row.description,
                "Original Price (PHP)": row.price,
                "Extraction Notes": row.extraction_notes,
            }
        )
    return pd.DataFrame(data, columns=RAW_COLUMNS)


def conflict_dataframe(rows: List[Dict[str, object]]) -> pd.DataFrame:
    return pd.DataFrame([{col: row.get(col, "") for col in CONFLICT_COLUMNS} for row in rows], columns=CONFLICT_COLUMNS)


def evidence_dataframe(rows: List[Dict[str, object]]) -> pd.DataFrame:
    return pd.DataFrame([{col: row.get(col, "") for col in EVIDENCE_COLUMNS} for row in rows], columns=EVIDENCE_COLUMNS)


# -----------------------------
# Excel export
# -----------------------------


def _write_df_to_sheet(wb: Workbook, title: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(title=title)
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, column in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=column)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    your_price_col = df.columns.get_loc("Your Price (PHP)") + 1 if "Your Price (PHP)" in df.columns else None
    original_price_col = df.columns.get_loc("Original Price (PHP)") + 1 if "Original Price (PHP)" in df.columns else None
    use_price_col = df.columns.get_loc("Use Price (PHP)") + 1 if "Use Price (PHP)" in df.columns else None

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for col_idx, column in enumerate(df.columns, start=1):
            value = row[column]
            if pd.isna(value):
                value = None
            if column == "Use Price (PHP)" and your_price_col and original_price_col:
                your_ref = f"{get_column_letter(your_price_col)}{row_idx}"
                original_ref = f"{get_column_letter(original_price_col)}{row_idx}"
                ws.cell(row=row_idx, column=col_idx, value=f'=IF({your_ref}="",{original_ref},{your_ref})')
            else:
                ws.cell(row=row_idx, column=col_idx, value=value)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    for col_idx, column in enumerate(df.columns, start=1):
        if column in {"Original Price (PHP)", "Your Price (PHP)", "Use Price (PHP)"}:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = '₱#,##0.00'
        if column == "Confidence":
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = '0%'

    for col_idx, column in enumerate(df.columns, start=1):
        max_len = len(str(column))
        for row_idx in range(2, min(ws.max_row, 250) + 1):
            max_len = max(max_len, len(_clean_text(ws.cell(row_idx, col_idx).value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 42)


def build_workbook(
    applications_df: pd.DataFrame,
    raw_df: pd.DataFrame,
    review_df: Optional[pd.DataFrame] = None,
    conflict_df: Optional[pd.DataFrame] = None,
    evidence_df: Optional[pd.DataFrame] = None,
) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    _write_df_to_sheet(wb, "Applications", applications_df)
    _write_df_to_sheet(wb, "Catalog Rows", raw_df)
    _write_df_to_sheet(wb, "Review Queue", review_df if review_df is not None else applications_df.iloc[0:0])
    _write_df_to_sheet(wb, "Conflict Log", conflict_df if conflict_df is not None else pd.DataFrame(columns=CONFLICT_COLUMNS))
    _write_df_to_sheet(wb, "Evidence Log", evidence_df if evidence_df is not None else pd.DataFrame(columns=EVIDENCE_COLUMNS))
    return wb


def workbook_to_bytes(wb: Workbook) -> bytes:
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_demo_excel(pdf_path: str | Path, output_path: str | Path) -> None:
    parsed = parse_catalog_pdf(pdf_path)
    applications_df = applications_dataframe(parsed.application_rows)
    raw_df = raw_rows_dataframe(parsed.raw_rows)
    review_df = pd.DataFrame(parsed.review_rows, columns=MAIN_COLUMNS)
    conflict_df = conflict_dataframe(parsed.conflict_rows)
    evidence_df = evidence_dataframe(parsed.evidence_rows)
    wb = build_workbook(applications_df, raw_df, review_df, conflict_df, evidence_df)
    wb.save(str(output_path))
